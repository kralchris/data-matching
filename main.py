"""
Implementation of a program that merges Table 1 with Table 2
into a combined Table #3 - based on "Model Name" from Table #1.

@author: Kristijan <kristijan.sarin@gmail.com>
"""


import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelDataProcessor:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.sheet = self.workbook.active

    def read_table(self, min_row, min_col, max_col, column_names):
        max_row = self.sheet.max_row
        for row in range(max_row, min_row - 1, -1):
            if any(self.sheet.cell(row=row, column=col).value is not None for col in range(min_col, max_col + 1)):
                break
        data = pd.DataFrame(self.sheet.iter_rows(
            values_only=True, min_row=min_row, min_col=min_col, max_col=max_col, max_row=row))
        data.columns = column_names
        return data

    def process_data(self):
        table_1 = self.read_table(
            14, 2, 4, ['Record ID 1', 'Vendor 1', 'Model Name 1'])
        table_1['Model Name 1 Lower'] = table_1['Model Name 1'].str.lower()

        table_2 = self.read_table(
            14, 7, 10, ['Record ID 2', 'Vendor 2', 'Description', 'Model Name 2'])
        table_2['Model Name 2 Lower'] = table_2['Model Name 2'].str.lower()

        merged_table = pd.merge(
            table_1, table_2, left_on='Model Name 1 Lower', right_on='Model Name 2 Lower', how='inner')
        merged_table.drop(
            columns=['Model Name 1 Lower', 'Model Name 2 Lower'], inplace=True)

        table_3 = merged_table[['Record ID 1', 'Vendor 1',
                                'Model Name 1', 'Description', 'Model Name 2', 'Record ID 2']]
        table_3.columns = ['Record ID (from Table #1)', 'Vendor (from Table #1)',
                           'Model Name (from Table #1)', 'Description (from Table #2)', 'Model Name', 'Record ID']
        self.write_back_to_excel(table_3)

    def write_back_to_excel(self, dataframe):
        for r_idx, row in enumerate(dataframe_to_rows(dataframe, index=False, header=False), 14):
            for c_idx, value in enumerate(row, 12):
                self.sheet.cell(row=r_idx, column=c_idx, value=value)
        self.workbook.save('data_output.xlsx')


if __name__ == "__main__":
    file_path = 'Data.xlsx'
    processor = ExcelDataProcessor(file_path)
    processor.process_data()
